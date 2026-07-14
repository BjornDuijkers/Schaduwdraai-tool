from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ComparisonResult, ParsedDocument
from .text_utils import money


STATUS_FILLS = {
    "OK": PatternFill("solid", fgColor="D9EAD3"),
    "VERSCHIL": PatternFill("solid", fgColor="F4CCCC"),
    "ALLEEN_IN_A": PatternFill("solid", fgColor="FCE5CD"),
    "ALLEEN_IN_B": PatternFill("solid", fgColor="CFE2F3"),
    "MATCH": PatternFill("solid", fgColor="D9EAD3"),
}


def create_excel_report(
    result: ComparisonResult,
    doc_a: ParsedDocument,
    doc_b: ParsedDocument,
    output_path: Path,
) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Samenvatting"

    _write_summary(summary_sheet, result, doc_a, doc_b)
    _write_employees(workbook.create_sheet("Medewerkers"), result)
    _write_components(workbook.create_sheet("Componenten"), result)
    _write_warnings(workbook.create_sheet("Waarschuwingen"), result)
    _write_mapping_tips(workbook.create_sheet("Mapping tips"), result)
    _write_issues(workbook.create_sheet("Issues"), result)

    for sheet in workbook.worksheets:
        _autosize(sheet)

    workbook.save(output_path)
    return output_path


def create_issues_report(
    result: ComparisonResult,
    doc_a: ParsedDocument,
    doc_b: ParsedDocument,
    output_path: Path,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Issues"
    _write_issues(sheet, result)
    _autosize(sheet)
    workbook.save(output_path)
    return output_path


def _write_summary(sheet, result: ComparisonResult, doc_a: ParsedDocument, doc_b: ParsedDocument) -> None:
    rows = [
        ("Document A", doc_a.source_name),
        ("Document B", doc_b.source_name),
        ("Periode A", doc_a.normalized_period or doc_a.period or ""),
        ("Periode B", doc_b.normalized_period or doc_b.period or ""),
        ("Provider A", doc_a.provider or ""),
        ("Provider B", doc_b.provider or ""),
        ("Scenario A", doc_a.scenario or ""),
        ("Scenario B", doc_b.scenario or ""),
        ("Loonstroken document A", result.summary["loonstroken_document_a"]),
        ("Loonstroken document B", result.summary["loonstroken_document_b"]),
        ("Uitgelezen componenten document A", result.summary["componenten_document_a"]),
        ("Uitgelezen componenten document B", result.summary["componenten_document_b"]),
        ("Gematchte medewerkers", result.summary["gematchte_medewerkers"]),
        ("Alleen in document A", result.summary["alleen_in_a"]),
        ("Alleen in document B", result.summary["alleen_in_b"]),
        ("Componentregels", result.summary["componentregels"]),
        ("Componenten OK", result.summary["componenten_ok"]),
        ("Componentverschillen", result.summary["componentverschillen"]),
        ("Componenten alleen in A", result.summary["componenten_alleen_in_a"]),
        ("Componenten alleen in B", result.summary["componenten_alleen_in_b"]),
        ("Waarschuwingen", result.summary["waarschuwingen"]),
    ]
    sheet.append(("Kenmerk", "Waarde"))
    for row in rows:
        sheet.append(row)
    _style_header(sheet)


def _write_employees(sheet, result: ComparisonResult) -> None:
    headers = [
        "Status",
        "Naam",
        "Geboortedatum",
        "Medewerkercode",
        "Pagina's A",
        "Pagina's B",
        "Opmerking",
    ]
    sheet.append(headers)
    for row in result.employees:
        sheet.append(
            [
                row.status,
                row.employee_name,
                row.birth_date,
                row.employee_code,
                row.source_a_pages,
                row.source_b_pages,
                row.match_note,
            ]
        )
    _style_table(sheet, status_column=1)


def _write_components(sheet, result: ComparisonResult) -> None:
    headers = [
        "Afwijking ID",
        "Status",
        "Reviewstatus",
        "Review opmerking",
        "Issue export",
        "Naam",
        "Geboortedatum",
        "Medewerkercode",
        "Component",
        "Component A",
        "Bedrag A",
        "Component B",
        "Bedrag B",
        "Verschil B-A",
        "Pagina's A",
        "Pagina's B",
    ]
    sheet.append(headers)
    for row in result.components:
        sheet.append(
            [
                row.deviation_id,
                row.status,
                row.review_status,
                row.review_note,
                "Ja" if row.export_issue else "Nee",
                row.employee_name,
                row.birth_date,
                row.employee_code,
                row.canonical_component,
                row.component_a,
                money(row.amount_a),
                row.component_b,
                money(row.amount_b),
                money(row.difference),
                row.pages_a,
                row.pages_b,
            ]
        )
    _style_table(sheet, status_column=2, money_columns=(11, 13, 14))


def _write_warnings(sheet, result: ComparisonResult) -> None:
    sheet.append(("Waarschuwing",))
    for warning in result.warnings:
        sheet.append((warning,))
    _style_table(sheet)


def _write_mapping_tips(sheet, result: ComparisonResult) -> None:
    sheet.append(("Component A", "Component B", "Status", "Advies"))
    for row in result.components:
        if row.status == "ALLEEN_IN_A":
            sheet.append((row.component_a, "", row.status, "Controleer of dit component in document B anders heet."))
        elif row.status == "ALLEEN_IN_B":
            sheet.append(("", row.component_b, row.status, "Controleer of dit component in document A anders heet."))
        elif row.status == "VERSCHIL":
            sheet.append((row.component_a, row.component_b, row.status, "Bedrag wijkt af."))
    _style_table(sheet, status_column=3)


def _write_issues(sheet, result: ComparisonResult) -> None:
    headers = [
        "Afwijking ID",
        "Reviewstatus",
        "Opmerking",
        "Naam",
        "Geboortedatum",
        "Medewerkercode",
        "Component",
        "Status",
        "Component A",
        "Bedrag A",
        "Component B",
        "Bedrag B",
        "Verschil B-A",
        "Pagina's A",
        "Pagina's B",
    ]
    sheet.append(headers)

    for row in result.components:
        if not row.export_issue and row.review_status != "exporteren als issue":
            continue
        sheet.append(
            [
                row.deviation_id,
                row.review_status,
                row.review_note,
                row.employee_name,
                row.birth_date,
                row.employee_code,
                row.canonical_component,
                row.status,
                row.component_a,
                money(row.amount_a),
                row.component_b,
                money(row.amount_b),
                money(row.difference),
                row.pages_a,
                row.pages_b,
            ]
        )
    _style_table(sheet, status_column=8, money_columns=(10, 12, 13))


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"


def _style_table(sheet, *, status_column: int | None = None, money_columns: tuple[int, ...] = ()) -> None:
    _style_header(sheet)
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        if status_column:
            status = row[status_column - 1].value
            fill = STATUS_FILLS.get(status)
            if fill:
                for cell in row:
                    cell.fill = fill
        for column_index in money_columns:
            if len(row) >= column_index:
                row[column_index - 1].number_format = '#,##0.00'


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        column = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column].width = min(max(max_length + 2, 12), 70)
